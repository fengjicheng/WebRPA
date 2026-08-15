# -*- coding: utf-8 -*-
"""回归：Excel 读取结果必须能被变量系统安全消费。

复现的线上缺陷（用户反馈）：
    工作流「读取区域」→「打印日志 {range_data}」执行失败，
    TypeError: Object of type ArrayFormula is not JSON serializable

实测确认的三个独立问题，本文件逐一锁定：
1. base.py 渲染 {变量} 时裸 json.dumps，任何非 JSON 原生类型都会把工作流带崩
   （ArrayFormula 只是其一，日期单元格返回的 datetime 更高频）
2. excel_read_range / excel_read_cell 曾硬编码 data_only=False，读到的是公式而不是值
3. 脏对象会被写进全局变量存储，污染下一次执行并让接口序列化失败

用真实 openpyxl 生成 xlsx，不做 mock。
"""
import json

import pytest

pytestmark = pytest.mark.regression


@pytest.fixture
def xlsx_含公式与日期(tmp_path):
    """生成含普通公式、数组公式、日期、纯数值的 xlsx（无缓存计算值）。"""
    import openpyxl
    from openpyxl.worksheet.formula import ArrayFormula
    from datetime import datetime

    path = tmp_path / "报表.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["A2"] = 2
    ws["B1"] = "=SUM(A1:A2)"
    ws["B2"] = ArrayFormula("B2", "=SUM(A1:A2*2)")
    ws["C1"] = datetime(2026, 8, 5, 22, 27, 5)
    ws["C2"] = datetime(2026, 8, 6)
    wb.save(path)
    wb.close()
    return str(path)


async def _read_range(path, ctx, **extra):
    from app.executors.advanced_openpyxl import ExcelReadRangeExecutor

    config = {"filePath": path, "range": "A1:C2", "resultVariable": "range_data"}
    config.update(extra)
    return await ExcelReadRangeExecutor().execute(config, ctx)


async def test_读取区域后打印日志不再崩溃(xlsx_含公式与日期, make_context):
    """缺陷主现场：读取区域 → resolve_value('{range_data}') 必须成功。"""
    ctx = make_context()
    result = await _read_range(xlsx_含公式与日期, ctx)
    assert result.success

    rendered = ctx.resolve_value("{range_data}")
    assert isinstance(rendered, str)
    # 渲染结果必须是合法 JSON，且不含 Python 对象的 repr 痕迹
    parsed = json.loads(rendered)
    assert len(parsed) == 2
    assert "ArrayFormula object" not in rendered


async def test_读取区域结果全是json原生类型(xlsx_含公式与日期, make_context):
    ctx = make_context()
    await _read_range(xlsx_含公式与日期, ctx)
    grid = ctx.get_variable("range_data")

    for row in grid:
        for cell in row:
            assert cell is None or isinstance(cell, (str, int, float, bool)), (
                f"单元格值类型未规范化: {type(cell).__name__}"
            )
    # 变量本身必须可直接序列化（全局变量存储、接口返回都依赖这一点）
    json.dumps(grid, ensure_ascii=False)


async def test_日期单元格规范化为文本(xlsx_含公式与日期, make_context):
    ctx = make_context()
    await _read_range(xlsx_含公式与日期, ctx)
    grid = ctx.get_variable("range_data")
    assert grid[0][2] == "2026-08-05 22:27:05"
    # 零点整的日期不带多余的 00:00:00
    assert grid[1][2] == "2026-08-06"


async def test_默认读值模式下无缓存公式回退为公式文本(xlsx_含公式与日期, make_context):
    """openpyxl 生成的文件没有公式缓存值，纯值模式会读成 None，必须回退给出公式文本。"""
    ctx = make_context()
    result = await _read_range(xlsx_含公式与日期, ctx)
    grid = ctx.get_variable("range_data")

    assert grid[0][1] == "=SUM(A1:A2)"
    assert grid[1][1] == "=SUM(A1:A2*2)"
    assert "无缓存计算值" in result.message


async def test_公式模式读到公式文本而非对象(xlsx_含公式与日期, make_context):
    ctx = make_context()
    result = await _read_range(xlsx_含公式与日期, ctx, readContent="formula")
    assert result.success
    grid = ctx.get_variable("range_data")
    assert grid[0][1] == "=SUM(A1:A2)"
    assert grid[1][1] == "=SUM(A1:A2*2)"
    json.dumps(grid, ensure_ascii=False)


async def test_值模式有缓存时读到计算结果(tmp_path, make_context):
    """普通数值单元格（等价于有缓存值的公式格）：值模式必须原样读到数字，且不带回退提示。"""
    import openpyxl

    path = tmp_path / "值文件.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = 3
    wb.save(path)
    wb.close()

    from app.executors.advanced_openpyxl import ExcelReadCellExecutor

    ctx = make_context()
    result = await ExcelReadCellExecutor().execute(
        {"filePath": str(path), "cell": "A1", "resultVariable": "v"}, ctx
    )
    assert result.success
    assert ctx.get_variable("v") == 3
    assert "无缓存计算值" not in result.message


async def test_读取单元格数组公式不再泄漏对象(xlsx_含公式与日期, make_context):
    from app.executors.advanced_openpyxl import ExcelReadCellExecutor

    ctx = make_context()
    result = await ExcelReadCellExecutor().execute(
        {"filePath": xlsx_含公式与日期, "cell": "B2", "resultVariable": "v"}, ctx
    )
    assert result.success
    assert ctx.get_variable("v") == "=SUM(A1:A2*2)"
    assert ctx.resolve_value("{v}") == "=SUM(A1:A2*2)"


async def test_读取excel模块日期不再崩(tmp_path, make_context):
    """advanced.py 的 read_excel 走 data_only=True，日期同样返回 datetime。"""
    import openpyxl
    from datetime import datetime

    from app.executors.advanced import ReadExcelExecutor

    path = tmp_path / "日期.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = datetime(2026, 8, 5, 22, 27)
    ws["A2"] = "文本"
    wb.save(path)
    wb.close()

    ctx = make_context()
    result = await ReadExcelExecutor().execute(
        {"fileName": str(path), "readMode": "range", "startCell": "A1", "endCell": "A2",
         "variableName": "v"}, ctx
    )
    assert result.success
    assert ctx.get_variable("v")[0][0] == "2026-08-05 22:27:00"
    json.loads(ctx.resolve_value("{v}"))


async def test_读取为字典数组日期不再崩(tmp_path, make_context):
    import openpyxl
    from datetime import datetime

    from app.executors.advanced_openpyxl_pro import ExcelReadDictsExecutor

    path = tmp_path / "记录.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["姓名", "入职日期"])
    ws.append(["张三", datetime(2026, 8, 5)])
    wb.save(path)
    wb.close()

    ctx = make_context()
    result = await ExcelReadDictsExecutor().execute(
        {"filePath": str(path), "resultVariable": "records"}, ctx
    )
    assert result.success
    assert ctx.get_variable("records")[0]["入职日期"] == "2026-08-05"
    json.loads(ctx.resolve_value("{records}"))


async def test_读取公式模块数组公式不再泄漏对象(xlsx_含公式与日期, make_context):
    from app.executors.advanced_openpyxl_pro import ExcelReadFormulaExecutor

    ctx = make_context()
    result = await ExcelReadFormulaExecutor().execute(
        {"filePath": xlsx_含公式与日期, "cell": "B2", "mode": "formula", "resultVariable": "f"}, ctx
    )
    assert result.success
    assert ctx.get_variable("f") == "=SUM(A1:A2*2)"


def test_变量渲染对任意非json类型都不抛异常(make_context):
    """不依赖 Excel 的通用契约：{变量} 渲染绝不能因为值的类型而失败。"""
    from datetime import datetime

    class 自定义对象:
        def __str__(self):
            return "对象文本"

    ctx = make_context()
    ctx.set_variable("行", [[datetime(2026, 8, 5, 1, 2, 3), 自定义对象()]])
    ctx.set_variable("单值", datetime(2026, 8, 5))

    rendered = ctx.resolve_value("{行}")
    assert json.loads(rendered) == [["2026-08-05 01:02:03", "对象文本"]]
    assert ctx.resolve_value("{单值}") == "2026-08-05"


def test_全局变量接口返回的载荷可序列化(client):
    """缺陷三的对外表现：接口必须能正常返回（历史上被路由遮蔽 + 序列化失败双重击穿）。

    run_execution 里「变量入库前规范化」这一步由
    tests/regression/test_regression_manual_path_events.py 的出口收尾用例覆盖（真实源码路径）。
    """
    from app.api import workflows

    原有变量 = dict(workflows.global_variables)
    try:
        workflows.global_variables["回归_数字"] = 1
        resp = client.get("/api/workflows/global-variables")
        assert resp.status_code == 200
        assert resp.json()["variables"]["回归_数字"] == 1
    finally:
        workflows.global_variables.clear()
        workflows.global_variables.update(原有变量)
